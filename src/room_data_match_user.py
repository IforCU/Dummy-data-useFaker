from openpyxl import load_workbook
import random

# 파일 경로 설정
rooms_file_path = 'rooms.xlsx'
users_file_path = 'user.xlsx'

# rooms 엑셀 파일 열기
rooms_wb = load_workbook(rooms_file_path)
rooms_ws = rooms_wb.active

# users 엑셀 파일 열기
users_wb = load_workbook(users_file_path)
users_ws = users_wb.active

# user 데이터에서 role이 HOST인 행의 user_id와 superhost 정보를 가져오기
hosts = []
for idx, row in enumerate(users_ws.iter_rows(min_row=2, values_only=True), start=2):
    role = row[2]  # role 컬럼 인덱스 (0-based 인덱스: 2)
    if role == "HOST":
        user_id = idx  # user_id로 행 번호 사용 (엑셀은 1-based 인덱스)
        superhost = row[3]  # superhost 컬럼 인덱스 (0-based 인덱스: 3)
        hosts.append((user_id, superhost))

# rooms 엑셀 파일의 각 행에 대해 user_id 및 guest_favorite 업데이트
for row in rooms_ws.iter_rows(min_row=2, values_only=False):
    # 랜덤한 host를 선택
    user_id, superhost = random.choice(hosts)

    # user_id 업데이트
    row[20].value = user_id  # user_id 컬럼 인덱스 (1-based 인덱스: 21)

    # guest_favorite 업데이트
    row[3].value = "TRUE" if superhost else "FALSE"  # guest_favorite 컬럼 인덱스 (1-based 인덱스: 4)

# 업데이트된 rooms 엑셀 파일 저장
rooms_wb.save('rooms_updated.xlsx')
