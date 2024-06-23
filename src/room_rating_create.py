import openpyxl
from collections import defaultdict

# 리뷰 파일 열기
review_wb = openpyxl.load_workbook('review.xlsx')
review_ws = review_wb.active

# 데이터 저장용 딕셔너리 생성
room_data = defaultdict(lambda: {
    'rating': 0,
    'review_count': 0,
    'cleanliness': 0,
    'accuracy': 0,
    'checkin': 0,
    'communication': 0,
    'location': 0,
    'value': 0
})

# 리뷰 데이터 읽기
for row in review_ws.iter_rows(min_row=2, max_col=10, values_only=True):
    room_id = row[9]  # J열 (room_id)
    cleanliness = row[1]  # B열
    accuracy = row[2]  # C열
    checkin = row[3]  # D열
    communication = row[4]  # E열
    location = row[5]  # F열
    value = row[6]  # G열

    # room_id에 대한 정보 업데이트
    room_data[room_id]['cleanliness'] += cleanliness
    room_data[room_id]['accuracy'] += accuracy
    room_data[room_id]['checkin'] += checkin
    room_data[room_id]['communication'] += communication
    room_data[room_id]['location'] += location
    room_data[room_id]['value'] += value
    room_data[room_id]['review_count'] += 1

# 평균 값 계산
for room_id, data in room_data.items():
    review_count = data['review_count']
    if review_count > 0:
        data['rating'] = (data['cleanliness'] + data['accuracy'] + data['checkin'] + data['communication'] + data['location'] + data['value']) / (6 * review_count)
        data['cleanliness'] /= review_count
        data['accuracy'] /= review_count
        data['checkin'] /= review_count
        data['communication'] /= review_count
        data['location'] /= review_count
        data['value'] /= review_count

# 새로운 엑셀 파일 생성
output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# 헤더 작성
header = ['room_id', 'rating', 'review_count', 'cleanliness', 'accuracy', 'checkin', 'communication', 'location', 'value']
output_ws.append(header)

# 데이터 작성
for room_id, data in room_data.items():
    row = [
        room_id,
        round(data['rating'], 2),
        data['review_count'],
        round(data['cleanliness'], 2),
        round(data['accuracy'], 2),
        round(data['checkin'], 2),
        round(data['communication'], 2),
        round(data['location'], 2),
        round(data['value'], 2)
    ]
    output_ws.append(row)

# 새로운 엑셀 파일 저장
output_wb.save('room_ratings.xlsx')
