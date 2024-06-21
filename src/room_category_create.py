import openpyxl

# 엑셀 파일을 읽어오기
room_wb = openpyxl.load_workbook('room.xlsx')
room_category_list_wb = openpyxl.load_workbook('room_category_list.xlsx')

# 각 파일의 첫 번째 시트를 가져오기
room_sheet = room_wb.active
room_category_list_sheet = room_category_list_wb.active

# 새로운 엑셀 파일 생성
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# 새 sheet의 헤더 설정
new_sheet['A1'] = 'room_id'
new_sheet['B1'] = 'room_category_list_id'

# 데이터 매핑 및 새로운 엑셀 파일 작성
new_row = 2
for room_row in range(2, room_sheet.max_row + 1):
    sub_title = room_sheet.cell(room_row, 2).value  # room 데이터의 두 번째 열(sub_title)
    room_id = room_row  # room_id는 해당 행의 row 값

    if sub_title:  # sub_title이 None이 아닌 경우에만 처리
        for category_row in range(2, room_category_list_sheet.max_row + 1):
            room_category = room_category_list_sheet.cell(category_row, 2).value  # room_category_list 데이터의 두 번째 열(room_category)

            if room_category and room_category in sub_title:
                room_category_list_id = room_category_list_sheet.cell(category_row, 1).value

                new_sheet.cell(new_row, 1, room_id)
                new_sheet.cell(new_row, 2, room_category_list_id)
                new_row += 1

# 새로운 엑셀 파일 저장
new_wb.save('new_room_data.xlsx')
