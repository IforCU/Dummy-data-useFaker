from openpyxl import load_workbook
import random

# rooms 엑셀 파일 열기
rooms_wb = load_workbook('rooms.xlsx')
rooms_ws = rooms_wb.active

# room_category_list 엑셀 파일 열기
category_wb = load_workbook('room_category_list.xlsx')
category_ws = category_wb.active

room_categories = [row[0] for row in category_ws.iter_rows(min_row=2, max_col=1, values_only=True)]

# rooms 엑셀 파일의 각 행에 대해 sub_title 작성 및 업데이트
for row in rooms_ws.iter_rows(min_row=2, values_only=False):
    city = row[10].value  # city 컬럼 인덱스 (1-based 인덱스: 11)
    country = row[9].value  # country 컬럼 인덱스 (1-based 인덱스: 10)
    random_category = random.choice(room_categories)  # room_category_list에서 랜덤하게 선택

    if city and country:
        sub_title = f"{city} {country} 의 {random_category}"
        row[1].value = sub_title  # sub_title 컬럼 인덱스 (1-based 인덱스: 2)

# 업데이트된 rooms 엑셀 파일 저장
rooms_wb.save('rooms_updated.xlsx')
