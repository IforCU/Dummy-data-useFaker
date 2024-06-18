import random
import openpyxl
from faker import Faker
import os

fake = Faker('ko_KR')

# 워크북과 시트를 생성합니다.
workbook = openpyxl.Workbook()
sheet = workbook.active

# 컬럼 설정
columns = ["id", "name", "image", "role", "superhost", "host_since", "created_at", "country", "location", "phone_number", "email"]
sheet.append(columns)

# 리뷰 파일 디렉토리 설정 (src 폴더 내의 py파일 기준으로 public 폴더 설정)
current_directory = os.path.dirname(os.path.abspath(__file__))
review_directory = os.path.join(current_directory, '..', 'public')
review_files = [f'rooms{i}.xlsx' for i in range(1, 32)]

print(f"Current directory: {current_directory}")
print(f"Review directory: {review_directory}")

user_id = 1

for review_file in review_files:
    file_path = os.path.join(review_directory, review_file)
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        continue  # 파일이 없으면 다음 파일로 넘어갑니다.
    
    # 리뷰 파일을 엽니다.
    review_wb = openpyxl.load_workbook(file_path)
    review_sheet = review_wb.active

    # 첫 번째 행이 있는지 확인합니다.
    rows = list(review_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not rows:
        print(f"No header found in {file_path}")
        continue

    # 첫 번째 행에서 컬럼 인덱스 찾기
    header = [cell for cell in rows[0]]
    try:
        name_idx = header.index("name")
        country_idx = header.index("country")
        image_idx = header.index("image")
    except ValueError as e:
        print(f"Error in {file_path}: {e}")
        continue

    # 각 리뷰 파일에서 데이터를 읽습니다.
    for row in review_sheet.iter_rows(min_row=2, values_only=True):
        name = row[name_idx]
        country = row[country_idx]
        image = row[image_idx]

        signup_date = fake.date_between(start_date='-20y', end_date='today')
        hosting_date = fake.date_between(start_date=signup_date, end_date='today') if random.choice([True, False]) else None
        role = "HOST" if hosting_date else "USER"
        superhost = fake.boolean() if role == "HOST" else False

        user_data = [
            user_id,  # ID
            name,  # 이름
            image,  # 이미지
            role,  # 역할
            superhost,  # 슈퍼호스트
            hosting_date.isoformat() if hosting_date else "N/A",  # 호스팅 날짜
            signup_date.isoformat(),  # 가입 날짜
            country,  # 국적
            fake.city(),  # 내 위치
            fake.phone_number(),  # 전화번호
            fake.email()  # 이메일
        ]
        sheet.append(user_data)
        user_id += 1

# 엑셀 파일 저장
output_file_path = os.path.join(current_directory, 'user_table.xlsx')
workbook.save(output_file_path)

print(f"User table saved at: {output_file_path}")
